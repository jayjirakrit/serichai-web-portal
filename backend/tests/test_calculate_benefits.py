import base64
import io
from datetime import date

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from services.accounts_service import (
    BENEFIT_UNFORMULATED_COLUMNS,
    OUTPUT_COLUMNS,
    SEVERANCE_MONTHS_TABLE,
    SURVIVAL_PROBABILITY_TABLE,
    BenefitsRequestError,
    _bracket_lookup,
    _datedif_years,
    _normalize_employee_id,
    build_exception_report,
    calculate_benefits,
    compute_benefit_rows,
    compute_carry_forward,
)

HEADERS = ["คำนำหน้า", "ชื่อ", "สกุล", "อัตราค่าแรง", "วันเดือนปีเกิด", "เริ่มทำงาน", "สถานะ"]
ID_HEADERS = ["รหัสพนักงาน"] + HEADERS
PREVIOUS_HEADERS = ["รหัสพนักงาน", "ผลประโยชน์ของพนักงานที่คาดว่าต้องจ่าย ณ วันสิ้นปีปัจจุบัน"]


def _master_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _master_bytes_with_id(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(ID_HEADERS)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _previous_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(PREVIOUS_HEADERS)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _find_row(output_rows: list[dict], first_name: str) -> dict:
    return next(r for r in output_rows if r["ชื่อ"] == first_name)


@pytest.fixture(autouse=True)
def fixed_cycle_year(monkeypatch):
    # Matches Employee_Benefit_Sample.xlsx's own reference "Date" (31-Dec BE2567),
    # so the expected values below can be checked against the Sample's known output.
    monkeypatch.setattr("services.accounts_service.current_cycle_be_year", lambda: 2567)


def test_datedif_years_not_yet_at_anniversary_subtracts_one():
    start = pd.Series([pd.Timestamp(2534, 8, 1)])
    end = pd.Series([pd.Timestamp(2543, 6, 5)])  # end's month/day precedes start's
    assert _datedif_years(start, end).tolist() == [8]


def test_datedif_years_past_anniversary_no_subtraction():
    start = pd.Series([pd.Timestamp(2560, 3, 2)])
    end = pd.Series([pd.Timestamp(2567, 12, 31)])
    assert _datedif_years(start, end).tolist() == [7]


def test_bracket_lookup_severance_months_exact_and_top_breakpoint():
    values = pd.Series([8, 20])
    assert _bracket_lookup(values, SEVERANCE_MONTHS_TABLE).tolist() == [8, 13.33]


def test_bracket_lookup_survival_probability_below_lowest_and_exact_breakpoint():
    # 55 is itself a breakpoint (start of the 55-60 bracket), so an exact
    # match returns that bracket's 100%, not the preceding 51-55 bracket's 90%.
    values = pd.Series([-5, 54, 55, 84])
    assert _bracket_lookup(values, SURVIVAL_PROBABILITY_TABLE).tolist() == [0.0, 0.9, 1.0, 1.0]


def test_no_proration_at_or_above_retirement_age_matches_sample_row_5():
    # Marked resigned: retirement-projection columns are only populated for
    # resigned rows now (accounts_service.py compute_benefit_rows).
    content = _master_bytes(
        [["นาง", "สวอง", "วายุพัฒน์", 11070, date(2483, 6, 5), date(2534, 8, 1), "ลาออก"]]
    )

    rows, exceptions, _ = compute_benefit_rows(content, "master.xlsx")

    assert exceptions == []
    row = _find_row(rows, "สวอง")
    assert row["อายุปัจจุบัน"] == 84
    assert row["อายุงานถึงปีเกษียณ"] == 8
    assert row["อายุงานถึงปีปัจจุบัน"] == 33
    assert row["อายุงานคงเหลือ"] == 0
    assert row["เงินเดือน ณ วันเกษียณ"] == 11070
    assert row["ผลประโยชน์ของพนักงานที่ต้องจ่าย ณ วันเกษียณ"] == 88560
    assert row["ความน่าจะเป็นในการอยู่จนถึงวันเกษียณ"] == 1
    assert row["ประมาณการหนี้สินผลประโยชน์พนักงานที่คาดว่าจะต้องจ่าย ณ วันเกษียณ"] == 88560
    assert row["ผลประโยชน์ของพนักงานที่คาดว่าต้องจ่าย ณ วันสิ้นปีปัจจุบัน"] == 88560


def test_proration_applies_below_retirement_age():
    # Marked resigned: retirement-projection columns are only populated for
    # resigned rows now (accounts_service.py compute_benefit_rows).
    content = _master_bytes(
        [["นาย", "ทดสอบ", "โปรเรต", 10000, date(2526, 11, 17), date(2560, 3, 2), "ลาออก"]]
    )

    rows, _, _ = compute_benefit_rows(content, "master.xlsx")

    row = _find_row(rows, "ทดสอบ")
    assert row["อายุปัจจุบัน"] == 41
    assert row["อายุงานถึงปีเกษียณ"] == 26
    assert row["อายุงานถึงปีปัจจุบัน"] == 7
    liability_at_retirement = row["ประมาณการหนี้สินผลประโยชน์พนักงานที่คาดว่าจะต้องจ่าย ณ วันเกษียณ"]
    expected = round(liability_at_retirement * 7 / 26, 2)
    assert row["ผลประโยชน์ของพนักงานที่คาดว่าต้องจ่าย ณ วันสิ้นปีปัจจุบัน"] == expected
    assert row["ผลประโยชน์ของพนักงานที่คาดว่าต้องจ่าย ณ วันสิ้นปีปัจจุบัน"] < liability_at_retirement


def test_employee_below_eligibility_age_silently_excluded():
    content = _master_bytes(
        [["นาย", "ยังไม่ถึง", "เกณฑ์", 10000, date(2540, 1, 1), date(2560, 3, 2), None]]
    )

    rows, exceptions, _ = compute_benefit_rows(content, "master.xlsx")

    assert rows == []
    assert exceptions == []  # not eligible yet is a normal state, not a data error


def test_row_missing_date_of_birth_excluded_and_reported():
    content = _master_bytes(
        [["นาย", "ไม่มี", "วันเกิด", 10000, None, date(2560, 3, 2), None]]
    )

    rows, exceptions, _ = compute_benefit_rows(content, "master.xlsx")

    assert rows == []
    assert len(exceptions) == 1
    assert exceptions[0]["category"] == "missingRequiredField"
    assert exceptions[0]["employee_name"] == "นายไม่มี วันเกิด"


def test_resigned_flag_set_and_unformulated_columns_left_blank():
    content = _master_bytes(
        [["นาง", "ลาออก", "แล้ว", 10000, date(2526, 11, 17), date(2560, 3, 2), "ลาออก"]]
    )

    rows, _, _ = compute_benefit_rows(content, "master.xlsx")

    row = _find_row(rows, "ลาออก")
    assert row["ลาออก"] == "ใช่"
    assert row["ยอดยกมา"] is None
    for col in BENEFIT_UNFORMULATED_COLUMNS:
        assert row[col] is None


def test_missing_required_master_column_raises_request_error():
    wb = Workbook()
    ws = wb.active
    ws.append(["คำนำหน้า", "ชื่อ", "สกุล", "อัตราค่าแรง"])  # no วันเดือนปีเกิด/เริ่มทำงาน
    ws.append(["นาย", "ไม่มี", "คอลัมน์", 10000])
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(BenefitsRequestError):
        compute_benefit_rows(buf.getvalue(), "master.xlsx")


def test_calculate_benefits_assembles_summary_and_files():
    content = _master_bytes(
        [
            ["นาง", "สวอง", "วายุพัฒน์", 11070, date(2483, 6, 5), date(2534, 8, 1), None],
            ["นาย", "ไม่มี", "วันเกิด", 10000, None, date(2560, 3, 2), None],
        ]
    )

    result = calculate_benefits(content, "master.xlsx")

    assert result["summary"]["total_employees_out"] == 1
    assert result["summary"]["computed_count"] == 1
    assert result["summary"]["exception_count"] == 1
    assert result["summary"]["exceptions_by_category"]["missingRequiredField"] == 1
    assert result["employee_benefits_report"]["filename"] == "employee_benefits_report.xlsx"
    assert result["exception_report"]["filename"] == "exception_report.xlsx"
    assert result["employee_benefits_report"]["content_base64"]
    assert result["exception_report"]["content_base64"]


def test_output_workbook_uses_template_and_appends_rows_after_header():
    content = _master_bytes(
        [["นาง", "สวอง", "วายุพัฒน์", 11070, date(2483, 6, 5), date(2534, 8, 1), None]]
    )

    result = calculate_benefits(content, "master.xlsx")
    out_bytes = base64.b64decode(result["employee_benefits_report"]["content_base64"])
    ws = load_workbook(io.BytesIO(out_bytes)).active

    # Rows 1-3 (company header, title/date/assumptions, instructions) and the
    # merged instructional ranges come straight from the template file —
    # unaffected by the calculation.
    assert ws.cell(1, 1).value == "บริษัท ช. ไพศาล จำกัด"
    assert ws.cell(3, 1).value == "กรอกข้อมูล"
    assert "A3:H3" in [str(r) for r in ws.merged_cells.ranges]

    # Row 4 is still the template's own header row, unchanged.
    assert [ws.cell(4, c + 1).value for c in range(len(OUTPUT_COLUMNS))] == OUTPUT_COLUMNS

    # The computed employee is appended directly below the header, not into
    # a freshly-built sheet.
    assert ws.cell(5, 1).value == 1
    assert ws.cell(5, 2).value is None  # no employee ID column in this master file
    assert ws.cell(5, 3).value == "นาง"
    assert ws.max_row == 5


# --- Previous-year benefit carry-forward (specs/002-previous-benefit-carryforward) ---


def test_normalize_employee_id_variants():
    assert _normalize_employee_id(1002.0) == "1002"
    assert _normalize_employee_id(" 1002 ") == "1002"
    assert _normalize_employee_id(float("nan")) is None
    assert _normalize_employee_id(None) is None
    assert _normalize_employee_id("") is None
    assert _normalize_employee_id("   ") is None


def test_compute_carry_forward_unique_match_builds_lookup_with_no_exceptions():
    content = _previous_bytes([["1002", 50000], ["1003", 60000]])

    lookup, exceptions = compute_carry_forward(content, "previous.xlsx")

    assert lookup == {"1002": 50000, "1003": 60000}
    assert exceptions == []


def test_compute_carry_forward_duplicate_id_excluded_and_reported():
    content = _previous_bytes([["1002", 50000], ["1002", 70000]])

    lookup, exceptions = compute_carry_forward(content, "previous.xlsx")

    assert "1002" not in lookup
    assert len(exceptions) == 1
    assert exceptions[0]["category"] == "duplicateEmployeeId"
    assert exceptions[0]["employee_id"] == "1002"
    assert exceptions[0]["employee_name"] is None


def test_compute_carry_forward_missing_required_column_raises():
    wb = Workbook()
    ws = wb.active
    ws.append(["รหัสพนักงาน"])  # missing the liability column
    ws.append(["1002"])
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(BenefitsRequestError):
        compute_carry_forward(buf.getvalue(), "previous.xlsx")


def test_compute_carry_forward_empty_file_raises():
    with pytest.raises(BenefitsRequestError):
        compute_carry_forward(b"", "previous.xlsx")


def test_compute_carry_forward_no_file_returns_empty():
    assert compute_carry_forward(None, None) == ({}, [])


def test_compute_benefit_rows_carry_forward_lookup_matched_unmatched_and_blank_id():
    content = _master_bytes_with_id(
        [
            ["1002", "นาง", "จับคู่", "ได้", 11070, date(2483, 6, 5), date(2534, 8, 1), None],
            ["9999", "นาง", "ไม่จับคู่", "เลย", 11070, date(2483, 6, 5), date(2534, 8, 1), None],
            [None, "นาง", "ไม่มี", "รหัส", 11070, date(2483, 6, 5), date(2534, 8, 1), None],
        ]
    )

    rows, exceptions, _ = compute_benefit_rows(content, "master.xlsx", {"1002": 50000})

    assert exceptions == []
    assert _find_row(rows, "จับคู่")["ยอดยกมา"] == 50000
    assert _find_row(rows, "ไม่จับคู่")["ยอดยกมา"] is None
    assert _find_row(rows, "ไม่มี")["ยอดยกมา"] is None


def test_calculate_benefits_end_to_end_with_carry_forward_and_duplicate():
    master_content = _master_bytes_with_id(
        [["1002", "นาง", "สวอง", "วายุพัฒน์", 11070, date(2483, 6, 5), date(2534, 8, 1), None]]
    )
    previous_content = _previous_bytes([["1002", 50000], ["1005", 1], ["1005", 2]])

    result = calculate_benefits(
        master_content,
        "master.xlsx",
        previous_content=previous_content,
        previous_filename="previous.xlsx",
    )

    assert result["summary"]["exceptions_by_category"]["duplicateEmployeeId"] == 1

    out_bytes = base64.b64decode(result["employee_benefits_report"]["content_base64"])
    ws = load_workbook(io.BytesIO(out_bytes)).active
    carry_forward_col = OUTPUT_COLUMNS.index("ยอดยกมา") + 1
    assert ws.cell(5, carry_forward_col).value == 50000
    employee_id_col = OUTPUT_COLUMNS.index("รหัสพนักงาน") + 1
    assert ws.cell(5, employee_id_col).value == "1002"


def test_calculate_benefits_without_previous_file_is_regression_safe():
    content = _master_bytes(
        [["นาง", "สวอง", "วายุพัฒน์", 11070, date(2483, 6, 5), date(2534, 8, 1), None]]
    )

    result = calculate_benefits(content, "master.xlsx")

    assert result["summary"]["exceptions_by_category"]["duplicateEmployeeId"] == 0
    out_bytes = base64.b64decode(result["employee_benefits_report"]["content_base64"])
    ws = load_workbook(io.BytesIO(out_bytes)).active
    carry_forward_col = OUTPUT_COLUMNS.index("ยอดยกมา") + 1
    assert ws.cell(5, carry_forward_col).value is None


def test_duplicate_employee_id_exception_shape_and_report_rendering():
    content = _previous_bytes([["1002", 50000], ["1002", 70000]])
    _, exceptions = compute_carry_forward(content, "previous.xlsx")

    assert exceptions[0]["employee_id"] == "1002"
    assert exceptions[0]["employee_name"] is None

    combined = exceptions + [
        {"category": "missingRequiredField", "employee_id": None, "employee_name": "ทดสอบ คน", "detail": "x"}
    ]
    report_bytes = build_exception_report(combined)
    ws = load_workbook(io.BytesIO(report_bytes)).active
    rows = [tuple(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    assert ("duplicateEmployeeId", "1002", None, exceptions[0]["detail"]) in rows
    assert ("missingRequiredField", None, "ทดสอบ คน", "x") in rows

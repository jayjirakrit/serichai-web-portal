import base64
import io
from datetime import date, datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook

from util import be_dates

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "data" / "Employee_Benefit_Template.xlsx"

RESIGNED_MARKER = "ลาออก"
ELIGIBILITY_AGE = 35
RETIREMENT_AGE = 60
SALARY_INCREASE_RATE = 0.03

# Severance-pay entitlement per Thai Labor Protection Act §118/§120:
# ascending (min_years_of_service, months_of_pay) breakpoints, approximate
# ("VLOOKUP TRUE") match — sourced from the Sample workbook's สูตร sheet,
# rows 4-10.
SEVERANCE_MONTHS_TABLE: list[tuple[float, float]] = [
    (0, 0),
    (0.33, 1),
    (1, 3),
    (3, 6),
    (6, 8),
    (10, 10),
    (20, 13.33),
]

# Company-defined (non-statutory) probability of remaining employed until
# retirement, by current age: ascending (min_age, probability) breakpoints,
# approximate match — สูตร sheet, rows 14-19.
SURVIVAL_PROBABILITY_TABLE: list[tuple[float, float]] = [
    (1, 0.0),
    (31, 0.2),
    (41, 0.5),
    (46, 0.6),
    (51, 0.9),
    (55, 1.0),
]

COLUMN_ALIASES: dict[str, list[str]] = {
    "id": ["รหัสพนักงาน"],
    "prefix": ["คำนำหน้า"],
    "first_name": ["ชื่อ"],
    "last_name": ["สกุล", "นามสกุล"],
    "date_of_birth": ["วันเดือนปีเกิด"],
    "start_date": ["เริ่มทำงาน", "วันที่เริ่มทำงาน"],
    "status": ["สถานะ"],
    "current_salary": ["เงินเดือน ณ สิ้นปีปัจจุบัน", "เงินเดือน", "อัตราค่าแรง"],
}

MASTER_REQUIRED_FIELDS = [
    "prefix",
    "first_name",
    "last_name",
    "date_of_birth",
    "start_date",
    "current_salary",
]

OUTPUT_COLUMNS = [
    "ลำดับ",
    "คำนำหน้า",
    "ชื่อ",
    "นามสกุล",
    "วันเดือนปีเกิด",
    "วันที่เริ่มทำงาน",
    "เงินเดือน ณ สิ้นปีปัจจุบัน",
    "วันที่เกษียณ (อายุครบ 60)",
    "อายุปัจจุบัน",
    "อายุงานถึงปีเกษียณ",
    "อายุงานถึงปีปัจจุบัน",
    "อายุงานคงเหลือ",
    "เงินเดือน ณ วันเกษียณ",
    "ผลประโยชน์ของพนักงานที่ต้องจ่าย ณ วันเกษียณ",
    "ความน่าจะเป็นในการอยู่จนถึงวันเกษียณ",
    "ประมาณการหนี้สินผลประโยชน์พนักงานที่คาดว่าจะต้องจ่าย ณ วันเกษียณ",
    "ผลประโยชน์ของพนักงานที่คาดว่าต้องจ่าย ณ วันสิ้นปีปัจจุบัน",
    "ยอดยกมา",
    "ต้นทุน",
    "คชจ.บริหาร",
    "คชจ.ขาย",
    "ลาออก",
    "อายุงาน (ปี/เดือน)",
    "อายุ (ปี/เดือน)",
]

# No formula exists anywhere in the reference workbook for these columns —
# they stay blank in the output (research.md #9, FR-022).
BENEFIT_UNFORMULATED_COLUMNS = [
    "ยอดยกมา",
    "ต้นทุน",
    "คชจ.บริหาร",
    "คชจ.ขาย",
    "อายุงาน (ปี/เดือน)",
    "อายุ (ปี/เดือน)",
]


class BenefitsRequestError(Exception):
    """Raised when the input file cannot be processed at all."""


# --- File reading ------------------------------------------------------


def _read_raw(content: bytes, filename: str) -> pd.DataFrame:
    if filename.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(content), header=None, dtype=str)
    return pd.read_excel(io.BytesIO(content), sheet_name=0, header=None, engine="openpyxl")


def read_table(content: bytes, filename: str) -> pd.DataFrame | None:
    try:
        raw = _read_raw(content, filename)
    except Exception:
        return None
    if raw.empty:
        return None

    header_row_idx = 0
    for i, row in raw.iterrows():
        cells = [str(c).strip() for c in row.tolist() if pd.notna(c)]
        if "คำนำหน้า" in cells:
            header_row_idx = i
            break

    header = [str(c).strip() if pd.notna(c) else "" for c in raw.iloc[header_row_idx].tolist()]
    data = raw.iloc[header_row_idx + 1 :].copy()
    data.columns = header
    data = data.dropna(how="all")
    if data.empty:
        return None
    return data.reset_index(drop=True)


def resolve_columns(df: pd.DataFrame) -> dict[str, str | None]:
    columns = list(df.columns)
    resolved: dict[str, str | None] = {}
    for logical_field, aliases in COLUMN_ALIASES.items():
        resolved[logical_field] = next((a for a in aliases if a in columns), None)
    return resolved


def build_exception_report(exceptions: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Exceptions"
    ws.append(["Category", "Employee ID", "Employee Name", "Detail"])
    for exc in sorted(exceptions, key=lambda e: (e["category"], e.get("employee_name") or "")):
        ws.append([exc["category"], exc.get("employee_id"), exc.get("employee_name"), exc["detail"]])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_output_workbook(rows: list[dict], year_end_date: date) -> bytes:
    """Loads `Employee_Benefit_Template.xlsx` as the base workbook — preserving
    its company header, instructions, and column-header row exactly — and
    appends one row per employee directly below that header row."""
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError as exc:
        raise BenefitsRequestError(f"Output template not found at {TEMPLATE_PATH}.") from exc
    ws = wb.active
    ws.cell(row=2, column=7, value=year_end_date)  # "Date" label's value cell (G2)
    for i, row in enumerate(rows, start=1):
        record = dict(row)
        record["ลำดับ"] = i
        ws.append([record.get(c) for c in OUTPUT_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def current_cycle_be_year() -> int:
    return datetime.now().year + 543


# --- Formula helpers ---------------------------------------------------


def _bracket_lookup(values: pd.Series, table: list[tuple[float, float]]) -> pd.Series:
    """Excel VLOOKUP(value, table, col, TRUE)-equivalent: approximate match
    on ascending lower-bound breakpoints, returning the paired value."""
    breakpoints = [t[0] for t in table]
    payouts = [t[1] for t in table]
    idx = np.searchsorted(breakpoints, values.fillna(-np.inf), side="right") - 1
    idx = np.clip(idx, 0, len(table) - 1)
    return pd.Series([payouts[i] for i in idx], index=values.index)


def _datedif_years(start: pd.Series, end: pd.Series) -> pd.Series:
    """Excel DATEDIF(start, end, "Y")-equivalent: completed whole years."""
    years = end.dt.year - start.dt.year
    not_yet_anniversary = (end.dt.month < start.dt.month) | (
        (end.dt.month == start.dt.month) & (end.dt.day < start.dt.day)
    )
    return years - not_yet_anniversary.astype(int)


# --- Pipeline ------------------------------------------------------------


def compute_benefit_rows(master_content: bytes, master_filename: str) -> tuple[list[dict], list[dict], date]:
    """Reads the employee master file and computes one output row per
    employee who has reached the eligibility age and has all required
    fields populated. Returns (output_rows, exceptions, year_end_date).
    """
    master_df = read_table(master_content, master_filename)
    if master_df is None:
        raise BenefitsRequestError("Employee master file is empty or unreadable.")

    cols = resolve_columns(master_df)
    missing = [f for f in MASTER_REQUIRED_FIELDS if cols.get(f) is None]
    if missing:
        raise BenefitsRequestError(f"Master file missing required columns: {', '.join(missing)}")

    df = pd.DataFrame(
        {
            "prefix": master_df[cols["prefix"]].astype(str).str.strip(),
            "first_name": master_df[cols["first_name"]].astype(str).str.strip(),
            "last_name": master_df[cols["last_name"]].astype(str).str.strip(),
            "date_of_birth": pd.to_datetime(master_df[cols["date_of_birth"]], errors="coerce"),
            "start_date": pd.to_datetime(master_df[cols["start_date"]], errors="coerce"),
            "current_salary": pd.to_numeric(master_df[cols["current_salary"]], errors="coerce"),
            "status_raw": master_df[cols["status"]] if cols.get("status") else None,
        }
    )
    df = df[df["first_name"].str.len() > 0].reset_index(drop=True)

    year_end_date = be_dates.benefit_year_end_date(current_cycle_be_year())
    valuation_date  = pd.Timestamp(year_end_date)

    # Validate required fields and collect exceptions for employees with missing data.
    exceptions: list[dict] = []
    valid = df["date_of_birth"].notna() & df["start_date"].notna() & df["current_salary"].notna()
    for _, row in df[~valid].iterrows():
        name = f"{row['prefix']}{row['first_name']} {row['last_name']}".strip()
        exceptions.append(
            {
                "category": "missingRequiredField",
                "employee_id": None,
                "employee_name": name or None,
                "detail": "Missing date of birth, start date, or wage rate — cannot compute benefit.",
            }
        )

    valid_employees = df[valid].copy().reset_index(drop=True)
    if valid_employees.empty:
        return [], exceptions, year_end_date

    valuation_date_series = pd.Series([valuation_date ] * len(valid_employees))
    valid_employees["current_age"] = _datedif_years(valid_employees["date_of_birth"], valuation_date_series)

    # Employees who haven't reached the eligibility age (35 by year-end) are excluded
    valid_employees = valid_employees[valid_employees["current_age"] >= ELIGIBILITY_AGE].reset_index(drop=True)
    if valid_employees.empty:
        return [], exceptions, year_end_date

    valuation_date_series = pd.Series([valuation_date ] * len(valid_employees))
    valid_employees["retirement_date"] = pd.to_datetime(valid_employees["date_of_birth"].apply(be_dates.retirement_date))
    valid_employees["years_to_retirement"] = _datedif_years(valid_employees["start_date"], valid_employees["retirement_date"])
    valid_employees["years_to_present"] = _datedif_years(valid_employees["start_date"], valuation_date_series)
    valid_employees["years_remaining"] = (RETIREMENT_AGE - valid_employees["current_age"]).clip(lower=0)

    valid_employees["salary_at_retirement"] = (
        valid_employees["current_salary"] * (1 + SALARY_INCREASE_RATE) ** valid_employees["years_remaining"]
    ).round(0)

    severance_months = _bracket_lookup(valid_employees["years_to_retirement"], SEVERANCE_MONTHS_TABLE)
    valid_employees["severance_at_retirement"] = valid_employees["salary_at_retirement"] * severance_months

    valid_employees["survival_probability"] = _bracket_lookup(valid_employees["current_age"], SURVIVAL_PROBABILITY_TABLE)
    valid_employees["liability_at_retirement"] = valid_employees["severance_at_retirement"] * valid_employees["survival_probability"]

    prorated = (valid_employees["liability_at_retirement"] * valid_employees["years_to_present"] / valid_employees["years_to_retirement"]).round(2)
    valid_employees["liability_at_year_end"] = np.where(
        valid_employees["current_age"] >= RETIREMENT_AGE,
        valid_employees["liability_at_retirement"],
        np.where(valid_employees["years_to_retirement"] > 0, prorated, valid_employees["liability_at_retirement"]),
    )

    valid_employees["is_resigned"] = valid_employees["status_raw"].fillna("").astype(str).str.strip() == RESIGNED_MARKER

    output_rows: list[dict] = []
    for i, row in valid_employees.iterrows():
        record = dict.fromkeys(OUTPUT_COLUMNS)
        record.update(
            {
                "ลำดับ": i + 1,
                "คำนำหน้า": row["prefix"],
                "ชื่อ": row["first_name"],
                "นามสกุล": row["last_name"],
                "วันเดือนปีเกิด": row["date_of_birth"].date(),
                "วันที่เริ่มทำงาน": row["start_date"].date(),
                "เงินเดือน ณ สิ้นปีปัจจุบัน": row["current_salary"],
                "วันที่เกษียณ (อายุครบ 60)": row["retirement_date"].date(),
                "อายุปัจจุบัน": int(row["current_age"]),
                "อายุงานถึงปีเกษียณ": int(row["years_to_retirement"]),
                "อายุงานถึงปีปัจจุบัน": int(row["years_to_present"]),
                "อายุงานคงเหลือ": int(row["years_remaining"]),
                "เงินเดือน ณ วันเกษียณ": row["salary_at_retirement"],
                "ผลประโยชน์ของพนักงานที่ต้องจ่าย ณ วันเกษียณ": row["severance_at_retirement"],
                "ความน่าจะเป็นในการอยู่จนถึงวันเกษียณ": row["survival_probability"],
                "ประมาณการหนี้สินผลประโยชน์พนักงานที่คาดว่าจะต้องจ่าย ณ วันเกษียณ": row["liability_at_retirement"],
                "ผลประโยชน์ของพนักงานที่คาดว่าต้องจ่าย ณ วันสิ้นปีปัจจุบัน": row["liability_at_year_end"],
                "ลาออก": bool(row["is_resigned"]),
            }
        )
        # BENEFIT_UNFORMULATED_COLUMNS stay None.
        output_rows.append(record)

    return output_rows, exceptions, year_end_date


def calculate_benefits(master_content: bytes, master_filename: str) -> dict:
    """Orchestrator: computes the benefit rows and assembles the API-facing
    summary plus the two output files."""
    output_rows, exceptions, year_end_date = compute_benefit_rows(master_content, master_filename)

    exceptions_by_category = {"missingRequiredField": 0}
    for e in exceptions:
        exceptions_by_category[e["category"]] = exceptions_by_category.get(e["category"], 0) + 1

    summary = {
        "total_employees_out": len(output_rows),
        "computed_count": len(output_rows),
        "resigned_flagged_count": sum(1 for r in output_rows if r["ลาออก"]),
        "exception_count": len(exceptions),
        "exceptions_by_category": exceptions_by_category,
    }

    workbook_bytes = build_output_workbook(output_rows, year_end_date)
    exception_bytes = build_exception_report(exceptions)

    return {
        "summary": summary,
        "exceptions": exceptions,
        "employee_benefits_report": {
            "filename": "employee_benefits_report.xlsx",
            "content_base64": base64.b64encode(workbook_bytes).decode("ascii"),
        },
        "exception_report": {
            "filename": "exception_report.xlsx",
            "content_base64": base64.b64encode(exception_bytes).decode("ascii"),
        },
    }
